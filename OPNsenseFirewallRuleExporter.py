import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

# Function to extract aliases from the XML
def extract_aliases(root):
    aliases = {}
    for alias in root.findall('.//alias'):
        name = alias.find('name').text
        content = alias.find('content').text if alias.find('content') is not None else ''
        aliases[name] = content  # Store only the name and content
    return aliases

# Function to extract interface descriptions from the XML
def get_interface_mapping(root):
    interface_mapping = {}
    # Loop through each interface in the XML to get its name and description
    for interface in root.findall('.//interfaces/*'):
        interface_name = interface.tag
        interface_descr = interface.find('descr').text if interface.find('descr') is not None else interface_name
        interface_mapping[interface_name] = interface_descr
    return interface_mapping

# Function to parse the XML and extract firewall rules
def parse_firewall_rules(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Get interface name to description mapping
    interface_mapping = get_interface_mapping(root)

    # Extract aliases
    aliases = extract_aliases(root)

    rules_by_interface = {}

    # Loop through each rule in the firewall section of the XML
    for rule in root.findall('.//rule'):
        # Determine if the rule is activated based on the <disabled> tag
        if rule.find('disabled') is not None and rule.find('disabled').text == '1':
            activated = 'No'
        else:
            activated = 'Yes'
        
        protocol = rule.find('protocol').text if rule.find('protocol') is not None else 'any'

        # Extract the action (block or pass) from the <type> tag
        action = rule.find('type').text if rule.find('type') is not None else 'pass'

        # Extract source details
        source = rule.find('source')
        source_ip = source.find('network').text if source is not None and source.find('network') is not None else 'any'
        source_port = source.find('port').text if source is not None and source.find('port') is not None else 'any'

        # Map the source IP to the interface description before applying the "!" logic
        source_ip = interface_mapping.get(source_ip, source_ip)

        # Check if <not>1</not> exists for source
        if source.find('not') is not None and source.find('not').text == '1':
            source_ip = '!' + source_ip

        # Extract destination details
        destination = rule.find('destination')
        if destination is not None:
            destination_ip = destination.find('network').text if destination.find('network') is not None else 'any'
            destination_port = destination.find('port').text if destination.find('port') is not None else 'any'
        else:
            destination_ip = 'any'
            destination_port = 'any'

        # Map the destination IP to the interface description before applying the "!" logic
        destination_ip = interface_mapping.get(destination_ip, destination_ip)

        # Check if <not>1</not> exists for destination
        if destination is not None and destination.find('not') is not None and destination.find('not').text == '1':
            destination_ip = '!' + destination_ip

        # Extract gateway and schedule, and replace "none" with "*"
        gateway = rule.find('gateway').text if rule.find('gateway') is not None else '*'
        schedule = rule.find('sched').text if rule.find('sched') is not None else '*'

        # Replace "none" values with "*"
        gateway = gateway if gateway != 'none' else '*'
        schedule = schedule if schedule != 'none' else '*'

        # Replace "none" or "unknown" in rule_number and rules with "*"
        rule_number = rule.find('tracker').text if rule.find('tracker') is not None else '*'
        rule_number = rule_number if rule_number != 'unknown' else '*'
        
        rules = rule.find('rule').text if rule.find('rule') is not None else '*'
        rules = rules if rules != 'none' else '*'

        description = rule.find('descr').text if rule.find('descr') is not None else ''
        
        # Get the interface for the rule and map it to its description
        interface = rule.find('interface').text if rule.find('interface') is not None else 'unknown'
        interface_description = interface_mapping.get(interface, interface)

        # Ensure interface_description is not None and rename "unknown" to "WAN"
        if interface_description is None or interface_description == 'unknown':
            interface_description = 'WAN'

        # Handle Floating rules that involve multiple interfaces
        if ',' in interface_description:
            interface_description = 'Floating'

        # Append the rule data to the appropriate interface in the dictionary
        if interface_description not in rules_by_interface:
            rules_by_interface[interface_description] = []
        rules_by_interface[interface_description].append([activated, action, protocol, source_ip, source_port, destination_ip, destination_port, gateway, schedule, rule_number, description, rules])

    return rules_by_interface, aliases

# Function to write data to Excel with separate sheets for each interface and apply table formatting
def write_to_excel(data_by_interface, aliases, output_file):
    # Create a workbook
    wb = openpyxl.Workbook()
    
    # Loop over each interface and its associated rules
    for interface_description, rules in data_by_interface.items():
        # Create a new sheet for each interface description
        if interface_description in wb.sheetnames:
            ws = wb[interface_description]
        else:
            ws = wb.create_sheet(title=interface_description)

        # Write headers, including the new "Action" column after "Aktiviert?"
        headers = ['Aktiviert?', 'Action', 'Protocol', 'Source', 'Port', 'Destination', 'Port', 'Gateway', 'Schedule', 'Number of Rule', 'Description', 'Rules']
        ws.append(headers)

        # Write each rule as a row and add alias comments
        for row_idx, row in enumerate(rules, start=2):  # Start from row 2 because row 1 is for headers
            ws.append(row)

            # Check and add comments for aliases in the relevant columns
            if row[3] in aliases:  # Source IP alias
                comment_text = f"Alias: {aliases[row[3]]}"
                ws.cell(row=row_idx, column=4).comment = Comment(comment_text, "Alias")

            if row[4] in aliases:  # Source Port alias
                comment_text = f"Alias: {aliases[row[4]]}"
                ws.cell(row=row_idx, column=5).comment = Comment(comment_text, "Alias")

            if row[5] in aliases:  # Destination IP alias
                comment_text = f"Alias: {aliases[row[5]]}"
                ws.cell(row=row_idx, column=6).comment = Comment(comment_text, "Alias")

            if row[6] in aliases:  # Destination Port alias
                comment_text = f"Alias: {aliases[row[6]]}"
                ws.cell(row=row_idx, column=7).comment = Comment(comment_text, "Alias")

        # Define the range for the table (including headers)
        table_range = f'A1:{chr(64+len(headers))}{len(rules)+1}'  # A1 to last column and last row

        # Create the table
        table = Table(displayName=f"Table_{interface_description}", ref=table_range)

        # Apply the "Blue, Table Medium 2" style
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style

        # Add the table to the sheet
        ws.add_table(table)
    
    # Remove the default sheet if no data was written to it
    if 'Sheet' in wb.sheetnames and not wb['Sheet'].max_row > 1:
        del wb['Sheet']

    # Save the workbook
    wb.save(output_file)
    print(f"Data written to {output_file}")

# Main function to execute the process
def main():
    xml_file = r'C:\GitHub\opnsense\config-opnsense.xml'  # Path to your XML file
    output_file = 'firewall_rules_by_interface.xlsx'  # Name of the output Excel file

    # Parse the XML and get the firewall rules by interface
    rules_by_interface, aliases = parse_firewall_rules(xml_file)

    # Write the data to Excel with separate sheets for each interface
    write_to_excel(rules_by_interface, aliases, output_file)

if __name__ == "__main__":
    main()
